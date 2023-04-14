from PyQt5 import QtWidgets, uic
from PyQt5.QtWidgets import *
from PyQt5.QtWidgets import QMessageBox
import sys
import sqlite3
from fpdf import FPDF
import os
from matplotlib import pyplot as plt
import time
from openpyxl import Workbook

banco = sqlite3.connect('databases/escola.db')
cursor = banco.cursor()

pdf = FPDF()
pdf.add_page()
pdf.set_font("Arial", size = 15)
pdf.set_text_color(233, 105, 105)

class Listar():
    def listardados():
        cursor.execute('SELECT * FROM escola')
        dados = cursor.fetchall()
        window.table.setRowCount(len(dados))
        window.table.setColumnCount(3)
        window.table.setHorizontalHeaderLabels(['Id', 'Nome', 'Qtd'])
        for i in range(0, len(dados)):
            for j in range(0, 3):
                window.table.setItem(i, j, QtWidgets.QTableWidgetItem(str(dados[i][j])))
        
    def listarqtd():
        cursor.execute(f"""SELECT * FROM escola ORDER BY Qtd DESC""")
        dados = cursor.fetchall()
        window.table.setRowCount(len(dados))
        window.table.setColumnCount(3)
        window.table.setHorizontalHeaderLabels(['Id', 'Nome', 'Qtd'])
        for i in range(0, len(dados)):
            for j in range(0, 3):
                window.table.setItem(i, j, QtWidgets.QTableWidgetItem(str(dados[i][j])))
        
    def listarqtdAsc():
        cursor.execute(f"""SELECT * FROM escola ORDER BY Qtd ASC""")
        dados = cursor.fetchall()
        window.table.setRowCount(len(dados))
        window.table.setColumnCount(3)
        window.table.setHorizontalHeaderLabels(['Id', 'Nome', 'Qtd'])
        for i in range(0, len(dados)):
            for j in range(0, 3):
                window.table.setItem(i, j, QtWidgets.QTableWidgetItem(str(dados[i][j])))
        
    def listarbusca():
        nome = window.nomeBuscarIn.text()
        cursor.execute(f"""SELECT Nome, Qtd,
                           CASE
                                WHEN Qtd < 3 THEN 'Doação de POBRE'
                                WHEN Qtd IN(4, 5) THEN 'Podia doar mais ne'
                                WHEN Qtd > 6 THEN 'Deus te abençoe'
                                ELSE 'É..'
                            END AS Avaliação
                            FROM escola
                            WHERE Nome = '{nome}'""")
        dados = cursor.fetchall()
        window.table.setRowCount(len(dados))
        window.table.setColumnCount(3)
        window.table.setHorizontalHeaderLabels(['Nome', 'Qtd', 'Avaliação'])
        for i in range(0, len(dados)):
            for j in range(0, 3):
                window.table.setItem(i, j, QtWidgets.QTableWidgetItem(str(dados[i][j])))
        
        cursor.execute(f"""SELECT SUM(Qtd) FROM escola WHERE nome = '{nome}'""")
    
        for a in cursor.fetchall():
            if str(a).replace('(', '',).replace(')', '').replace(',', '') == 'None':
                window.doado.setText(f'{nome} não consta no banco.')
            else:
                window.doado.setText(f'''R${str(a).replace('(', '',).replace(')', '').replace(',', '')}''')
        window.nomeBuscarIn.clear()
        
    def excluirAll():
         cursor.execute(F"""DELETE FROM escola""")
         window.idIn.clear()
         arre = window.arrecadado
         banco.commit()
            
         cursor.execute(f"""SELECT SUM(Qtd) FROM escola""")
         for a in cursor.fetchall():
             arre.setText(f'''R${str(a).replace('(', '',).replace(')', '').replace(',', '')}''')
                
         cursor.execute(f"""SELECT round(AVG(Qtd), 1) FROM escola""")
         for b in cursor.fetchall():
             window.media.setText(f'''R${str(b).replace('(', '',).replace(')', '').replace(',', '')}''')
            
         cursor.execute(f"""SELECT ROUND(SUM (Qtd) / 20 * 1, 3) AS MEDIA FROM escola""")
         for c in cursor.fetchall():
             window.por.setText(f'''{str(c).replace('(', '',).replace(')', '').replace(',', '')}%''')
         cursor.execute("""SELECT COUNT(Qtd) as total FROM escola""")
         for d in cursor.fetchall():
             window.total.setText(f'''{str(d).replace('(', '',).replace(')', '').replace(',', '')}''')
                
         Listar.listardados()
    
    def msg_clicked():
         msgBox = QMessageBox()
         msgBox.setIcon(QMessageBox.Critical)
         msgBox.setText("Certeza que deseja excluir tudo?")
         msgBox.setWindowTitle("Excluir tudo")
         msgBox.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        
         returnValue = msgBox.exec()
         if returnValue == QMessageBox.Ok:
             Listar.excluirAll()
         else:
            pass
    
def salvar():
    nome = window.nomeIn.text()
    qtd = window.qtdIn.value()
    arre = window.arrecadado
    banco = sqlite3.connect('databases/escola.db')
    cursor = banco.cursor()
    cursor.execute(F"""INSERT INTO escola VALUES(Null, '{nome}', '{qtd}')""")
    cursor.execute(f"""SELECT SUM(Qtd) FROM escola""")
    window.nomeIn.clear()
    banco.commit()
    for a in cursor.fetchall():
        texto = str(a).replace('(', '',).replace(')', '').replace(',', '')
        arre.setText(f'''R${texto}''')
        
    cursor.execute(f"""SELECT round(AVG(Qtd), 1) FROM escola""")
    for b in cursor.fetchall():
        window.media.setText(f'''R${str(b).replace('(', '',).replace(')', '').replace(',', '')}''')
    
    cursor.execute(f"""SELECT ROUND(SUM (Qtd) / 20 * 1, 3) AS MEDIA FROM escola""")
    for c in cursor.fetchall():
        window.por.setText(f'''{str(c).replace('(', '',).replace(')', '').replace(',', '')}%''')
    cursor.execute("""SELECT COUNT(Qtd) as total FROM escola""")
    for d in cursor.fetchall():
        window.total.setText(f'''{str(d).replace('(', '',).replace(')', '').replace(',', '')}''')
    Listar.listardados()
    
def excluir():
        id = window.idIn.text()
        cursor.execute(F"""DELETE FROM escola WHERE Id = '{id}'""")
        window.idIn.clear()
        arre = window.arrecadado
        banco.commit()
        
        cursor.execute(f"""SELECT SUM(Qtd) FROM escola""")
        for a in cursor.fetchall():
            arre.setText(f'''R${str(a).replace('(', '',).replace(')', '').replace(',', '')}''')
            
        cursor.execute(f"""SELECT round(AVG(Qtd), 1) FROM escola""")
        for b in cursor.fetchall():
            window.media.setText(f'''R${str(b).replace('(', '',).replace(')', '').replace(',', '')}''')
        
        cursor.execute(f"""SELECT ROUND(SUM (Qtd) / 20 * 1, 3) AS MEDIA FROM escola""")
        for c in cursor.fetchall():
            window.por.setText(f'''{str(c).replace('(', '',).replace(')', '').replace(',', '')}%''')
        cursor.execute("""SELECT COUNT(Qtd) as total FROM escola""")
        for d in cursor.fetchall():
            window.total.setText(f'''{str(d).replace('(', '',).replace(')', '').replace(',', '')}''')
            
        Listar.listardados()
       
def gPdf():
    pdf.cell(200, 10, txt = "Arrecadação Terceiro B", 
         ln = 1, align = 'C')
    
    pdf.cell(200, 10, txt = "Meta: R$2000", 
         ln = 1, align = 'L')
    
    cursor.execute(f"""SELECT SUM(Qtd) FROM escola""")
    for a in cursor.fetchall():
        pdf.cell(200, 10, txt = f"Arrecadado: R${str(a).replace('(', '',).replace(')', '').replace(',', '')}",
            ln = 2, align = 'L')
          
    cursor.execute(f"""SELECT round(AVG(Qtd), 1) FROM escola""")
    for b in cursor.fetchall():
        pdf.cell(200, 10, txt = f"Média: R${str(b).replace('(', '',).replace(')', '').replace(',', '')}",
            ln = 2, align = 'L')
          
    cursor.execute("""SELECT COUNT(Qtd) as total FROM escola""")
    for c in cursor.fetchall():
        pdf.cell(200, 10, txt = f"Total de Doações: {str(c).replace('(', '',).replace(')', '').replace(',', '')}",
            ln = 2, align = 'L')

    pdf.output("exported\GFG.pdf") 

    time.sleep(0.4)
    os.startfile("exported\GFG.pdf")
    
def gGrafico():
    cursor.execute("SELECT Qtd FROM escola GROUP BY Id")
    lista = []
    b = cursor.fetchall()
    for a in b:
        lista.append(float(str(a).replace('(', '',).replace(')', '').replace(',', '')))
        s = list(range(len(lista)))
        plt.title("Arrecadação 3b")
        plt.plot(s, lista, color='blue')
        plt.xlabel("Ordem das doações")
        plt.ylabel("Quantidade por doação")
        plt.legend(['Doações'])
        plt.show()

def gExcel():
    book = Workbook()
    sheet = book.active

    sheet['A1'] = 'ID'
    sheet['B1'] = 'Nome'
    sheet['C1'] = 'Valor'
    sheet['E1'] = 'Arrecadado'
    sheet['F1'] = 'Média'
    sheet['G1'] = 'Total'
    
    id = 2
    cursor.execute("SELECT ID FROM escola")
    for text in cursor.fetchall():
        sheet[f'A{id}'] = str(text).replace('(', '').replace(')', '').replace(',', '')
        id += 1
        
    nome = 2
    cursor.execute("SELECT Nome FROM escola")
    for text in cursor.fetchall():
        sheet[f'B{nome}'] = str(text).replace('(', '').replace(')', '').replace(',', '')
        nome += 1
        
    qtd = 2
    cursor.execute("SELECT Qtd FROM escola")
    for text in cursor.fetchall():
        sheet[f'C{qtd}'] = str(text).replace('(', '').replace(')', '').replace(',', '')
        qtd += 1
    
    cursor.execute("SELECT SUM(Qtd) FROM escola")
    for text in cursor.fetchall():
        sheet['E2'] = str(text).replace('(', '').replace(')', '').replace(',', '')
    
    cursor.execute("SELECT ROUND(AVG(Qtd), 2) FROM escola")
    for text in cursor.fetchall():
        sheet['F2'] = str(text).replace('(', '').replace(')', '').replace(',', '')
        
    cursor.execute("SELECT COUNT(Qtd) FROM escola")
    for text in cursor.fetchall():
        sheet['G2'] = str(text).replace('(', '').replace(')', '').replace(',', '')


    book.save("exported\escola.xlsx")
    os.popen("exported\escola.xlsx")

# Window
app = QtWidgets.QApplication(sys.argv)
window = uic.loadUi('interfaces/escola1.ui')
window.table.horizontalHeader().setStretchLastSection(True)
Listar.listardados()

# Style
window.table.verticalHeader().hide()

cursor.execute(f"""SELECT round(AVG(Qtd), 1) FROM escola""")
for a in cursor.fetchall():
    window.media.setText(f'''R${str(a).replace('(', '',).replace(')', '').replace(',', '')}''')
    
def pageCadastro():
    window.Pages.setCurrentWidget(window.page)
    window.cadastroPage.setStyleSheet('QPushButton' 
                                      '{'
                                      'border-radius: 10px;'
                                      'background-color: #8FA2DA;'
                                      'color: white;'
                                      '}')
    window.excluirPage.setStyleSheet('QPushButton'
                                       '{'
                                       'border-radius: 10px;'
                                       'background-color: #4FCFB4;'
                                       'color: white;'
                                       '}')
    window.buscarPage.setStyleSheet('QPushButton'
                                       '{'
                                       'border-radius: 10px;'
                                       'background-color: #4FCFB4;'
                                       'color: white;'
                                       '}')
    cursor.execute('SELECT * FROM escola')
    dados = cursor.fetchall()
    window.table.setRowCount(len(dados))
    window.table.setColumnCount(3)
    window.table.setHorizontalHeaderLabels(['Id', 'Nome', 'Qtd'])
    for i in range(0, len(dados)):
        for j in range(0, 3):
            window.table.setItem(i, j, QtWidgets.QTableWidgetItem(str(dados[i][j])))
            
def pageExcluir():
    window.cadastroPage.setStyleSheet('QPushButton'
                                       '{'
                                       'border-radius: 10px;'
                                       'background-color: #4FCFB4;'
                                       'color: white;'
                                       '}')
    window.buscarPage.setStyleSheet('QPushButton'
                                       '{'
                                       'border-radius: 10px;'
                                       'background-color: #4FCFB4;'
                                       'color: white;'
                                       '}')
    window.excluirPage.setStyleSheet('QPushButton' 
                                      '{'
                                      'border-radius: 10px;'
                                      'background-color: #8FA2DA;'
                                      'color: white;'
                                      '}')
    window.Pages.setCurrentWidget(window.page_2)
    cursor.execute('SELECT * FROM escola')
    dados = cursor.fetchall()
    window.table.setRowCount(len(dados))
    window.table.setColumnCount(3)
    window.table.setHorizontalHeaderLabels(['Id', 'Nome', 'Qtd'])
    for i in range(0, len(dados)):
        for j in range(0, 3):
            window.table.setItem(i, j, QtWidgets.QTableWidgetItem(str(dados[i][j])))
            
def pageBuscar():
    window.Pages.setCurrentWidget(window.page_3)
    window.buscarPage.setStyleSheet('QPushButton' 
                                      '{'
                                      'border-radius: 10px;'
                                      'background-color: #8FA2DA;'
                                      'color: white;'
                                      '}')
    window.excluirPage.setStyleSheet('QPushButton'
                                       '{'
                                       'border-radius: 10px;'
                                       'background-color: #4FCFB4;'
                                       'color: white;'
                                       '}')
    window.cadastroPage.setStyleSheet('QPushButton'
                                       '{'
                                       'border-radius: 10px;'
                                       'background-color: #4FCFB4;'
                                       'color: white;'
                                       '}')

# Botões
window.cadastroPage.clicked.connect(pageCadastro)
window.excluirPage.clicked.connect(pageExcluir)
window.buscarPage.clicked.connect(pageBuscar)

window.salvar.clicked.connect(salvar)
window.excluir.clicked.connect(excluir)
window.buscar.clicked.connect(Listar.listarbusca)
window.desc.clicked.connect(Listar.listarqtd)
window.asc.clicked.connect(Listar.listarqtdAsc)
window.organizar.clicked.connect(Listar.listardados)
window.gerarPdf.clicked.connect(gPdf)
window.gerarGrafico.clicked.connect(gGrafico)
window.gerarExcel.clicked.connect(gExcel)
window.excluirTudo.clicked.connect(Listar.msg_clicked)

# Banco
arre = window.arrecadado
cursor.execute(f"""SELECT SUM(Qtd) FROM escola""")
banco.commit()

for a in cursor.fetchall():
    arre.setText(f'''R${str(a).replace('(', '',).replace(')', '').replace(',', '')}''')
    
cursor.execute("""SELECT ROUND(SUM (Qtd) / 20 * 1, 3) AS MEDIA FROM escola""")
for b in cursor.fetchall():
    window.por.setText(f'''{str(b).replace('(', '',).replace(')', '').replace(',', '')}%''')
    
cursor.execute("""SELECT COUNT(Qtd) as total FROM escola""")
for c in cursor.fetchall():
    window.total.setText(f'''{str(c).replace('(', '',).replace(')', '').replace(',', '')}''')

window.show()
app.exec()